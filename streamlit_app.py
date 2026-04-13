import re
from io import BytesIO
from datetime import datetime

import pandas as pd
import streamlit as st
from pypdf import PdfReader
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FILL_COLOR = "1F4E78"
HEADER_FONT_COLOR = "FFFFFF"

MAJOR_TERMS = [
    "accounting", "business", "finance", "marketing", "management",
    "economics", "computer information systems", "information systems",
    "computer science", "education", "elementary education",
    "special education", "nursing", "engineering", "biology",
    "chemistry", "physics", "mathematics", "math", "psychology",
    "sociology", "social work", "criminal justice", "history",
    "english", "journalism", "communication", "communications",
    "agriculture", "mba", "master of business administration", "stem",
    "technology", "hospitality"
]

MONTHS_PATTERN = (
    r"January|February|March|April|May|June|July|August|September|October|November|December|"
    r"Jan|Feb|Mar|Apr|Jun|Jul|Aug|Sep|Sept|Oct|Nov|Dec"
)

STATE_WORDS = [
    "Illinois", "Indiana", "Kentucky", "Missouri", "Iowa", "United States"
]

COUNTY_WORDS = [
    "Coles County", "Cumberland County", "Clay County", "Richland County",
    "McLean County", "Cook County"
]

CITY_WORDS = [
    "Chicago", "Bloomington", "Normal", "Mattoon", "Charleston"
]


def clean_text(text: str) -> str:
    if not text:
        return ""
    text = text.replace("\r", "\n")
    text = re.sub(r"https?://\S+", "", text)
    text = re.sub(r"\n[ \t]*\n+", "\n\n", text)
    text = re.sub(r"[ \t]+", " ", text)
    return text.strip()


def normalize(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip() if text else ""


def safe(value):
    if value is None:
        return "Not Specified"
    if isinstance(value, str) and not value.strip():
        return "Not Specified"
    return value


def yes_no_or_not_specified(text, pattern):
    if not text or not text.strip():
        return "Not Specified"
    return "Yes" if re.search(pattern, text, re.I) else "No"


def unique_keep_order(items):
    seen = set()
    output = []
    for item in items:
        key = normalize(item).casefold()
        if key and key not in seen:
            seen.add(key)
            output.append(normalize(item))
    return output


def get_texts_from_upload(uploaded_file):
    uploaded_file.seek(0)
    reader = PdfReader(uploaded_file)
    raw_pages = [(page.extract_text() or "") for page in reader.pages]
    raw_text = "\n".join(raw_pages)
    cleaned_text = clean_text(raw_text)
    return raw_text, cleaned_text


def between(text, start, end):
    match = re.search(start + r"(.*?)" + end, text, re.S | re.I)
    return clean_text(match.group(1)) if match else ""


def single(text, label):
    patterns = [
        rf"{re.escape(label)}\s*:?\s*(.*?)(?=\n[A-Z][A-Za-z0-9 /&()'-]+(?:\n|$))",
        rf"{re.escape(label)}\s*:?\s*(.*?)(?=\s+[A-Z][A-Za-z0-9 /&()'-]+(?:\s|$))",
    ]
    stop_words = (
        r"\b(?:Visibility|Financial Information|Opportunity-Specific Information|"
        r"Department|Donor|Fund Code|Auxiliary Fund Code|Project ID|Type|"
        r"Post-Acceptance|Source|Visible Award Amount|Description|Full Description|"
        r"Keywords|Award Information|Questions)\b"
    )

    for pattern in patterns:
        match = re.search(pattern, text, re.S)
        if match:
            value = clean_text(match.group(1))
            value = re.split(stop_words, value)[0].strip(" :-")
            if value:
                return value
    return ""


def number(text, label):
    match = re.search(rf"{re.escape(label)}\s*:?\s*\$?([\d,]+(?:\.\d+)?)", text, re.I)
    return float(match.group(1).replace(",", "")) if match else None


def clean_for_requirement_matching(text):
    text = re.sub(r"\bEastern Illinois University\b", "", text, flags=re.I)
    text = re.sub(r"\bEIU\b", "", text, flags=re.I)
    return text


def extract_name(raw_text, fallback_name):
    lines = [line.strip() for line in raw_text.splitlines() if line.strip()]

    skip_exact = {
        "Portfolio", "Applicant", "Basic Information", "Financial Information",
        "Opportunity-Specific Information", "Award Information", "Questions",
        "Description", "Full Description", "Keywords"
    }

    skip_patterns = [
        r"^Fall\s+\d{4}$",
        r"^Spring\s+\d{4}$",
        r"^\|\s*Ended",
        r"^https?://",
        r"^\d+/\d+$",
        r"^\d{1,2}/\d{1,2}/\d{2,4},?",
        r"^Name\s+",
        r"^Deadline\s+",
        r"^Type\s+",
        r"^Department\s+",
        r"^Donor\s+",
        r"^Fund Code\s+",
    ]

    for line in lines:
        clean_line = normalize(line)
        if not clean_line or clean_line in skip_exact:
            continue
        if "Eastern Illinois University Scholarships" in clean_line:
            continue
        if any(re.search(pattern, clean_line, re.I) for pattern in skip_patterns):
            continue
        if 3 <= len(clean_line) <= 120:
            return clean_line

    return fallback_name.rsplit(".", 1)[0]


def extract_description_blocks(text):
    description = between(text, r"Description\s*:?\s*", r"\s*Full\s+[Dd]escription")
    full_description = between(
        text,
        r"Full\s+[Dd]escription\s*:?\s*",
        r"\s*(?:Keywords|Type|Department|Donor|Fund Code|$)"
    )
    if not description and not full_description:
        description = between(text, r"Description\s*:?\s*", r"\s*(?:Type|Department|Donor|Fund Code|$)")
    return clean_text(description), clean_text(full_description)


def find_requirement_sections(text):
    text = clean_text(text)
    section_labels = [
        "Eligibility", "Eligible", "Criteria", "Requirements", "Qualifications",
        "Selection Criteria", "Applicant Criteria", "Minimum Qualifications"
    ]

    sections = []
    for label in section_labels:
        pattern = rf"{label}\s*:?\s*(.*?)(?=\n[A-Z][A-Za-z0-9 /&()'-]{{2,40}}:?\s|\Z)"
        for match in re.finditer(pattern, text, re.S | re.I):
            chunk = clean_text(match.group(1))
            if chunk and len(chunk) > 20:
                sections.append(chunk)

    return unique_keep_order(sections)


def build_requirement_text(cleaned_text):
    description, full_description = extract_description_blocks(cleaned_text)
    found_sections = find_requirement_sections(cleaned_text)

    candidates = []
    if found_sections:
        candidates.extend(found_sections)
    if full_description:
        candidates.append(full_description)
    if description:
        candidates.append(description)

    return clean_for_requirement_matching(clean_text("\n\n".join(unique_keep_order(candidates))))


def split_sentences(text):
    if not text:
        return []
    text = normalize(text)
    return [s.strip() for s in re.split(r"(?<=[.!?])\s+|\n+", text) if s.strip()]


def parse_date_string(date_str):
    if not date_str:
        return ""

    date_str = normalize(date_str).rstrip(".,;:")
    date_str = re.sub(r"(\d)(st|nd|rd|th)\b", r"\1", date_str, flags=re.I)

    formats = [
        "%B %d, %Y", "%b %d, %Y",
        "%B %d %Y", "%b %d %Y",
        "%B %d", "%b %d",
        "%m/%d/%Y", "%m/%d/%y",
        "%m-%d-%Y", "%m-%d-%y",
        "%m/%d", "%m-%d",
    ]

    for fmt in formats:
        try:
            dt = datetime.strptime(date_str, fmt)
            if "%Y" not in fmt and "%y" not in fmt:
                dt = dt.replace(year=datetime.now().year)
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            continue

    return date_str


def extract_date_from_sentence(sentence):
    patterns = [
        rf"\b(({MONTHS_PATTERN})\.?\s+\d{{1,2}}(?:st|nd|rd|th)?(?:,\s*\d{{4}})?)\b",
        rf"\b(({MONTHS_PATTERN})\.?\s+\d{{1,2}}(?:st|nd|rd|th)?)\b",
        r"\b(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})\b",
        r"\b(\d{1,2}[/-]\d{1,2})\b",
    ]

    for pattern in patterns:
        match = re.search(pattern, sentence, re.I)
        if match:
            return parse_date_string(match.group(1))
    return ""


def deadline(text):
    if not text:
        return ""

    sentences = split_sentences(text)

    priority_patterns = [
        r"deadline",
        r"application deadline",
        r"applications due",
        r"due date",
        r"submit by",
        r"must be submitted by",
        r"apply by",
        r"application due",
        r"completed by",
        r"must apply by",
        r"must submit by",
        r"received by",
        r"no later than",
        r"priority consideration",
    ]

    for sentence in sentences:
        lowered = sentence.lower()
        if any(re.search(pattern, lowered) for pattern in priority_patterns):
            date_val = extract_date_from_sentence(sentence)
            if date_val:
                return date_val

    return ""


def gpa(text):
    patterns = [
        r"minimum GPA(?: of)?\s*(\d\.\d+|\d)",
        r"GPA(?: of)? at least\s*(\d\.\d+|\d)",
        r"cumulative GPA(?: of)?\s*(\d\.\d+|\d)",
        r"maintain a GPA(?: of)? at least\s*(\d\.\d+|\d)",
        r"have a GPA of\s*(\d\.\d+|\d)\s*or higher",
        r"must have (?:a )?minimum cumulative GPA of\s*(\d\.\d+|\d)",
        r"must maintain (?:a )?GPA of\s*(\d\.\d+|\d)",
        r"must have (?:a )?GPA of\s*(\d\.\d+|\d)",
        r"overall GPA(?: of)?\s*(\d\.\d+|\d)",
        r"grade point average(?: of)?\s*(\d\.\d+|\d)",
        r"minimum cumulative grade point average(?: of)?\s*(\d\.\d+|\d)",
        r"academic average(?: of)?\s*(\d\.\d+|\d)",
        r"scholastic average(?: of)?\s*(\d\.\d+|\d)",
        r"\b(\d\.\d+)\s*GPA\b",
        r"\bGPA\s*[:\-]?\s*(\d\.\d+|\d)\b",
    ]

    for pattern in patterns:
        match = re.search(pattern, text, re.I)
        if match:
            try:
                value = float(match.group(1))
                if 0.0 <= value <= 4.0:
                    return value
            except ValueError:
                continue

    if re.search(r"\bB average\b", text, re.I):
        return 3.0

    return None


def class_levels(text):
    levels = []

    checks = [
        ("Freshman", r"\bfreshm[ae]n?\b|first[- ]year"),
        ("Sophomore", r"\bsophomores?\b|second[- ]year"),
        ("Junior", r"\bjuniors?\b|third[- ]year"),
        ("Senior", r"\bseniors?\b|fourth[- ]year"),
        ("Graduate", r"\bgraduate\b|\bmaster'?s\b|\bMBA\b"),
        ("Undergraduate", r"\bundergraduate\b")
    ]

    for label, pattern in checks:
        if re.search(pattern, text, re.I):
            levels.append(label)

    return "; ".join(unique_keep_order(levels))


def normalize_location_phrase(text):
    text = normalize(text)
    text = text.strip(" .,:;")
    replacements = {
        "Mclean County": "McLean County",
        "Coles county": "Coles County",
        "Clay county": "Clay County",
        "Richland county": "Richland County",
        "Cumberland county": "Cumberland County",
        "Cook county": "Cook County",
        "United states": "United States",
    }
    for old, new in replacements.items():
        text = re.sub(rf"\b{re.escape(old)}\b", new, text, flags=re.I)
    return text


def clean_location_fragment(fragment):
    fragment = normalize_location_phrase(fragment)
    fragment = re.sub(
        r"\b(other|area|local|eligible|incoming|current|majoring|majors?|students?|applicants?)\b.*$",
        "",
        fragment,
        flags=re.I
    )
    fragment = fragment.strip(" ,;:-")
    return fragment


def geographic_preference(text):
    text = clean_for_requirement_matching(text)
    found = []

    phrase_patterns = [
        r"(?:must be|shall be|is open to|restricted to)\s+(?:a\s+)?resident of\s+([A-Za-z ,'/\-]+?)(?=[.;\n]| and | who | with |$)",
        r"(?:residents?|resident students?)\s+of\s+([A-Za-z ,'/\-]+?)(?=[.;\n]| and | who | with |$)",
        r"(?:students?|applicants?)\s+from\s+([A-Za-z ,'/\-]+?)(?=[.;\n]| and | who | with |$)",
        r"(?:must live|must reside|live|reside)\s+in\s+([A-Za-z ,'/\-]+?)(?=[.;\n]| and | who | with |$)",
        r"(?:permanent address|legal residence)\s+in\s+([A-Za-z ,'/\-]+?)(?=[.;\n]| and | who | with |$)",
        r"(?:graduates?|graduates? of students?)\s+of\s+high schools?\s+in\s+([A-Za-z ,'/\-]+?)(?=[.;\n]| and | who | with |$)",
        r"(?:preference(?: is)? given to|preference(?: will be)? given to)\s+(?:students?|applicants?)\s+from\s+([A-Za-z ,'/\-]+?)(?=[.;\n]| and | who | with |$)",
    ]

    for pattern in phrase_patterns:
        for match in re.finditer(pattern, text, re.I):
            candidate = clean_location_fragment(match.group(1))
            if 2 <= len(candidate) <= 60:
                found.append(candidate)

    requirement_windows = re.findall(
        r"((?:resident|reside|live|from|address|preference given|high school in)[^.:\n]{0,160})",
        text,
        re.I
    )

    for window in requirement_windows:
        for county in COUNTY_WORDS:
            if re.search(rf"\b{re.escape(county)}\b", window, re.I):
                found.append(county)
        for state in STATE_WORDS:
            if re.search(rf"\b{re.escape(state)}\b", window, re.I):
                found.append(state)
        for city in CITY_WORDS:
            if re.search(rf"\b{re.escape(city)}\b", window, re.I):
                found.append(city)

    high_school_match = re.search(r"\b([A-Z][a-z]+(?:\s+[A-Z][a-z]+)*)\s+High School\b", text)
    if high_school_match:
        found.append(f"{high_school_match.group(1)} High School")

    cleaned = []
    for item in unique_keep_order(found):
        item = clean_location_fragment(item)
        if not item:
            continue
        if len(item.split()) > 6:
            continue
        cleaned.append(item)

    return "; ".join(unique_keep_order(cleaned))


def financial_need(text):
    pattern = r"financial need|demonstrated need|FAFSA|Pell(?: Grant)?|need-based|economic need"
    return yes_no_or_not_specified(text, pattern)


def underserved_flag(text):
    if not text or not text.strip():
        return "Not Specified"

    strong_patterns = [
        r"low[- ]income",
        r"underprivileged",
        r"underserved",
        r"economically disadvantaged",
        r"financially disadvantaged",
        r"disadvantaged background",
        r"first[- ]generation",
        r"first generation",
        r"underrepresented",
        r"historically marginalized",
        r"financial hardship",
        r"economic hardship",
        r"limited financial resources",
        r"limited means",
        r"background of hardship",
        r"pell(?: grant)? eligible",
        r"pell recipient",
    ]

    moderate_patterns = [
        r"demonstrated financial need",
        r"high financial need",
        r"significant financial need",
        r"unmet financial need",
        r"financial need",
        r"economic need",
        r"need[- ]based",
        r"fafsa",
    ]

    for pattern in strong_patterns:
        if re.search(pattern, text, re.I):
            return "Yes"

    if re.search(
        r"(preference given|given preference|priority given|eligibility|eligible|selection|awarded to|must demonstrate|required to demonstrate)",
        text,
        re.I
    ):
        for pattern in moderate_patterns:
            if re.search(pattern, text, re.I):
                return "Yes"

    return "No"


def major_field(text):
    text = clean_for_requirement_matching(text)
    found = []

    context_patterns = [
        r"(?:major(?:ing)? in|majors? in|students? majoring in|degree in|pursuing a degree in|field of study in)\s+([^.:\n;]+)",
        r"(?:open to|restricted to)\s+([^.:\n;]+?)\s+majors?",
    ]

    for pattern in context_patterns:
        for match in re.finditer(pattern, text, re.I):
            chunk = match.group(1)
            for term in MAJOR_TERMS:
                if re.search(rf"\b{re.escape(term)}\b", chunk, re.I):
                    found.append(term)

    direct_student_patterns = [
        r"\b(accounting|business|education|finance|marketing|management|economics|engineering|nursing|psychology|history|journalism|communication|communications|technology|hospitality)\s+students\b",
        r"\bMBA students\b",
        r"\bSTEM students\b",
    ]

    for pattern in direct_student_patterns:
        for match in re.finditer(pattern, text, re.I):
            found.append(match.group(1))

    cleaned = []
    for item in unique_keep_order(found):
        low = item.lower()
        if low == "mba":
            cleaned.append("MBA")
        elif low == "stem":
            cleaned.append("STEM")
        else:
            cleaned.append(item.title())

    return ", ".join(unique_keep_order(cleaned))


def extract(uploaded_file):
    raw_text, text = get_texts_from_upload(uploaded_file)

    scholarship_name = extract_name(raw_text, uploaded_file.name)
    description, full_description = extract_description_blocks(text)
    requirement_text = build_requirement_text(text)

    broad_text = clean_text("\n\n".join([
        description,
        full_description,
        requirement_text,
        text
    ]))

    min_gpa = gpa(requirement_text) or gpa(broad_text)
    geo_pref = geographic_preference(requirement_text) or geographic_preference(broad_text)
    due_date = deadline(text) or deadline(requirement_text) or deadline(broad_text)

    low_income_flag = underserved_flag(requirement_text)
    if low_income_flag == "No":
        low_income_flag = underserved_flag(broad_text)

    essay_required = yes_no_or_not_specified(
        requirement_text,
        r"\bessay\b|brief summary|short essay|personal statement|statement of purpose"
    )
    recommendation_required = yes_no_or_not_specified(
        requirement_text,
        r"recommendation|reference letter|letter of recommendation"
    )
    leadership_mentioned = yes_no_or_not_specified(
        requirement_text,
        r"character|leadership|work ethic|personal values|motivation|goals|service|integrity"
    )

    return {
        "Scholarship Name": safe(scholarship_name),
        "Import Date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Fund Period Amount": number(text, "Fund Period Amount"),
        "Department": safe(single(raw_text, "Department")),
        "Donor": safe(single(raw_text, "Donor")),
        "Fund Code": safe(single(raw_text, "Fund Code")),
        "Opportunity Type": safe(single(raw_text, "Type")),
        "Post-Acceptance Enabled": safe(single(raw_text, "Post-Acceptance")),
        "Minimum GPA": safe(min_gpa),
        "Full-Time Required": yes_no_or_not_specified(requirement_text, r"full[- ]time"),
        "Class Level Eligible": safe(class_levels(requirement_text)),
        "Geographic Preference": safe(geo_pref),
        "Financial Need Considered": financial_need(requirement_text),
        "Low Income / Underprivileged Background": low_income_flag,
        "Major / Field of Study": safe(major_field(requirement_text)),
        "Deadline": safe(due_date),
        "Renewable / Reapply Allowed": yes_no_or_not_specified(
            requirement_text,
            r"apply again|eligible to apply again|continues to meet criteria|renewable|may reapply"
        ),
        "Essay Required": essay_required,
        "Recommendation Required": recommendation_required,
        "Character / Leadership Mentioned": leadership_mentioned,
        "Notes": "",
    }


def summary(df):
    rows = [
        ["Total Scholarships", len(df)],
        ["Total Fund Amount", pd.to_numeric(df["Fund Period Amount"], errors="coerce").fillna(0).sum()],
        ["", ""],
        ["Opportunity Type", "Count"]
    ]

    counts = df["Opportunity Type"].fillna("Unknown").replace("", "Unknown").value_counts()
    for key, value in counts.items():
        rows.append([key, int(value)])

    return pd.DataFrame(rows, columns=["Metric", "Value"])


def style_header_row(ws):
    header_fill = PatternFill(fill_type="solid", fgColor=HEADER_FILL_COLOR)
    header_font = Font(bold=True, color=HEADER_FONT_COLOR)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws.row_dimensions[1].height = 24


def format_scholarships_sheet(ws, df_columns):
    style_header_row(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    col_index = {name: idx + 1 for idx, name in enumerate(df_columns)}

    widths = {
        "Scholarship Name": 38,
        "Import Date": 21,
        "Fund Period Amount": 16,
        "Department": 24,
        "Donor": 24,
        "Fund Code": 16,
        "Opportunity Type": 18,
        "Post-Acceptance Enabled": 20,
        "Minimum GPA": 14,
        "Full-Time Required": 16,
        "Class Level Eligible": 24,
        "Geographic Preference": 32,
        "Financial Need Considered": 22,
        "Low Income / Underprivileged Background": 30,
        "Major / Field of Study": 26,
        "Deadline": 16,
        "Renewable / Reapply Allowed": 22,
        "Essay Required": 16,
        "Recommendation Required": 20,
        "Character / Leadership Mentioned": 28,
        "Notes": 28,
    }

    for col_name, width in widths.items():
        if col_name in col_index:
            ws.column_dimensions[get_column_letter(col_index[col_name])].width = width


def format_summary_sheet(ws):
    style_header_row(ws)
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18


def build_excel_bytes(df):
    summary_df = summary(df)
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Scholarships", index=False)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

        wb = writer.book
        format_scholarships_sheet(wb["Scholarships"], list(df.columns))
        format_summary_sheet(wb["Summary"])

    output.seek(0)
    return output


st.set_page_config(page_title="Scholarship Data Extraction Tool", layout="wide")
st.title("Scholarship Data Extraction Tool")
st.markdown("""
Upload scholarship PDFs, review the extracted fields, and download a formatted Excel workbook.

**How to use**
1. Upload one or more PDF files  
2. Click **Process PDFs**  
3. Download the finished Excel file
""")

uploaded_files = st.file_uploader("Upload PDF files", type=["pdf"], accept_multiple_files=True)

if uploaded_files:
    st.write(f"{len(uploaded_files)} PDF file(s) ready.")

    if st.button("Process PDFs"):
        data = []
        errors = []

        with st.spinner("Reading PDFs and building workbook..."):
            for f in uploaded_files:
                try:
                    data.append(extract(f))
                except Exception as e:
                    errors.append((f.name, str(e)))

        if data:
            df = pd.DataFrame(data)
            excel_file = build_excel_bytes(df)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"scholarship_database_v6_{timestamp}.xlsx"

            st.success("Processing complete.")
            st.dataframe(df, use_container_width=True)

            st.download_button(
                label="Download Excel file",
                data=excel_file,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        if errors:
            st.warning("Some files could not be processed.")
            for name, err in errors:
                st.error(f"{name}: {err}")
else:
    st.info("Upload one or more PDF files to begin.")
